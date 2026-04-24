import json
import os
import shutil
import smtplib
import subprocess
import sys
import tempfile
from datetime import datetime, timedelta, timezone
from email.utils import parseaddr
from email.message import EmailMessage
from io import BytesIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from argon2 import PasswordHasher
from argon2.exceptions import InvalidHash, VerifyMismatchError
from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from sqlalchemy import Boolean, Integer, String, Text, desc, func, inspect, or_, text
from werkzeug.utils import secure_filename


def resolve_timezone():
    try:
        return ZoneInfo("America/Argentina/Buenos_Aires")
    except ZoneInfoNotFoundError:
        return timezone(timedelta(hours=-3))


TZ_GMT_MINUS_3 = resolve_timezone()
ACTION_TYPES = {"BLOQUEAR", "DESBLOQUEAR", "NOVEDAD", "INCIDENCIA"}
ROLE_TYPES = {"ADMIN", "JRT", "SUPERVISOR", "LIDER", "GERENTE", "AUDITOR"}
PERMISSION_KEYS = {
    "can_access_admin_panel",
    "can_view_records",
    "can_search_records",
    "can_export_excel",
    "can_unlock_requests",
    "can_manage_users",
    "can_view_audit",
    "can_manage_policy",
    "can_view_password_health",
    "can_manage_permissions",
}
DEFAULT_ROLE_PERMISSIONS = {
    "ADMIN": {
        "can_access_admin_panel": True,
        "can_view_records": True,
        "can_search_records": True,
        "can_export_excel": True,
        "can_unlock_requests": True,
        "can_manage_users": True,
        "can_view_audit": True,
        "can_manage_policy": True,
        "can_view_password_health": True,
        "can_manage_permissions": True,
    },
    "JRT": {
        "can_access_admin_panel": True,
        "can_view_records": True,
        "can_search_records": True,
        "can_export_excel": True,
        "can_unlock_requests": True,
        "can_manage_users": False,
        "can_view_audit": False,
        "can_manage_policy": False,
        "can_view_password_health": False,
        "can_manage_permissions": False,
    },
    "SUPERVISOR": {
        "can_access_admin_panel": True,
        "can_view_records": True,
        "can_search_records": True,
        "can_export_excel": True,
        "can_unlock_requests": True,
        "can_manage_users": False,
        "can_view_audit": False,
        "can_manage_policy": False,
        "can_view_password_health": False,
        "can_manage_permissions": False,
    },
    "LIDER": {
        "can_access_admin_panel": True,
        "can_view_records": True,
        "can_search_records": True,
        "can_export_excel": False,
        "can_unlock_requests": True,
        "can_manage_users": False,
        "can_view_audit": False,
        "can_manage_policy": False,
        "can_view_password_health": False,
        "can_manage_permissions": False,
    },
    "GERENTE": {
        "can_access_admin_panel": True,
        "can_view_records": True,
        "can_search_records": True,
        "can_export_excel": True,
        "can_unlock_requests": False,
        "can_manage_users": False,
        "can_view_audit": True,
        "can_manage_policy": False,
        "can_view_password_health": True,
        "can_manage_permissions": False,
    },
    "AUDITOR": {
        "can_access_admin_panel": True,
        "can_view_records": True,
        "can_search_records": True,
        "can_export_excel": True,
        "can_unlock_requests": False,
        "can_manage_users": False,
        "can_view_audit": True,
        "can_manage_policy": False,
        "can_view_password_health": False,
        "can_manage_permissions": False,
    },
}
ARGON2_SCHEME = "argon2id"
ARGON2_MEMORY_COST = 19456
ARGON2_TIME_COST = 3
ARGON2_PARALLELISM = 1
DEFAULT_PASSWORD_POLICY = {
    "min_length": 8,
    "require_uppercase": True,
    "require_lowercase": True,
    "require_digit": True,
    "require_symbol": False,
    "expires_days": 90,
}
DEFAULT_NOTIFICATION_SETTINGS = {
    "enabled": False,
    "smtp_host": "",
    "smtp_port": 587,
    "smtp_username": "",
    "smtp_password": "",
    "smtp_use_tls": True,
    "smtp_use_ssl": False,
    "from_name": "Registro de Novedades",
    "from_email": "",
    "to_addresses": "",
    "cc_addresses": "",
    "subject_bloquear": "[Bloqueo] Nueva solicitud registrada",
    "subject_desbloquear": "[Desbloqueo] Nueva solicitud registrada",
    "subject_novedad": "[Novedad] Nueva solicitud registrada",
    "subject_incidencia": "[Incidencia] Nueva solicitud registrada",
}

def get_runtime_base_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.abspath(os.path.dirname(__file__))


def get_bundle_base_dir() -> str:
    return getattr(sys, "_MEIPASS", os.path.abspath(os.path.dirname(__file__)))


RUNTIME_BASE_DIR = get_runtime_base_dir()
BUNDLE_BASE_DIR = get_bundle_base_dir()
INSTANCE_DIR = os.path.join(RUNTIME_BASE_DIR, "instance")
ENV_DIR = os.path.join(RUNTIME_BASE_DIR, ".env")
REFERENCE_DATA_FILE = os.getenv("REFERENCE_DATA_FILE", os.path.join(RUNTIME_BASE_DIR, "data.xlsx"))
BUNDLED_REFERENCE_DATA_FILE = os.path.join(BUNDLE_BASE_DIR, "data.xlsx")
EVIDENCE_DIR = os.path.join(INSTANCE_DIR, "evidences")

os.makedirs(INSTANCE_DIR, exist_ok=True)
os.makedirs(ENV_DIR, exist_ok=True)
os.makedirs(EVIDENCE_DIR, exist_ok=True)


def ensure_reference_file_exists():
    if os.path.exists(REFERENCE_DATA_FILE):
        return
    if os.path.exists(BUNDLED_REFERENCE_DATA_FILE):
        try:
            shutil.copyfile(BUNDLED_REFERENCE_DATA_FILE, REFERENCE_DATA_FILE)
        except OSError:
            pass


ensure_reference_file_exists()
AUTH_PEPPER_FILE = os.getenv("AUTH_PEPPER_FILE", os.path.join(ENV_DIR, "auth_pepper.env"))
REFERENCE_CACHE = {
    "mtime": None,
    "bases": [],
    "transport_by_plate": {},
    "unit_type_by_plate": {},
}
ALLOWED_EVIDENCE_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".webp",
    ".pdf",
    ".txt",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
}
MAX_EVIDENCE_BYTES = 10 * 1024 * 1024


def clone_default_permissions() -> dict:
    return {
        role: {permission: bool(value) for permission, value in permissions.items()}
        for role, permissions in DEFAULT_ROLE_PERMISSIONS.items()
    }


def read_env_value(file_path: str, key: str) -> str:
    if not os.path.exists(file_path):
        return ""

    try:
        with open(file_path, "r", encoding="utf-8") as file:
            for line in file:
                line_clean = line.strip()
                if not line_clean or line_clean.startswith("#") or "=" not in line_clean:
                    continue
                env_key, env_value = line_clean.split("=", 1)
                if env_key.strip().lstrip("\ufeff") == key:
                    return env_value.strip().strip('"').strip("'")
    except OSError:
        return ""

    return ""


AUTH_PEPPER = os.getenv("AUTH_PEPPER", "") or read_env_value(AUTH_PEPPER_FILE, "AUTH_PEPPER")
LEGACY_ONLY_ADMIN = os.getenv("LEGACY_ONLY_ADMIN", "0").strip().lower() in {"1", "true", "si", "yes", "on"}
password_hasher = PasswordHasher(
    time_cost=ARGON2_TIME_COST,
    memory_cost=ARGON2_MEMORY_COST,
    parallelism=ARGON2_PARALLELISM,
)
SESSION_USERNAME_KEY = "session_username"
SESSION_ROLE_KEY = "session_role"
SESSION_SOURCE_KEY = "session_source"

app = Flask(
    __name__,
    template_folder=os.path.join(BUNDLE_BASE_DIR, "templates"),
    static_folder=os.path.join(BUNDLE_BASE_DIR, "static"),
)
default_sqlite_path = os.path.join(INSTANCE_DIR, "bloqueos.db")
LEGACY_ADMIN_HASH_FILE = os.path.join(RUNTIME_BASE_DIR, "admin_key.hash")
ADMIN_HASH_FILE = os.getenv("ADMIN_HASH_FILE", os.path.join(ENV_DIR, "admin_key.hash"))
app.secret_key = os.getenv("APP_SECRET_KEY", "") or os.getenv("FLASK_SECRET_KEY", "") or os.urandom(32).hex()
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv(
    "DATABASE_URL",
    f"sqlite:///{default_sqlite_path}",
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


class BlockRequest(db.Model):
    __tablename__ = "block_requests"

    id = db.Column(db.Integer, primary_key=True)
    created_at_system = db.Column(db.DateTime(timezone=True), nullable=False)
    event_datetime = db.Column(db.DateTime(timezone=True), nullable=False)

    action_type = db.Column(db.String(20), nullable=False, default="BLOQUEAR")
    related_request_id = db.Column(db.Integer, nullable=True)
    authorized_by_user = db.Column(db.String(120), nullable=True)

    carga_nro = db.Column(db.String(50), nullable=True)
    patente_primaria = db.Column(db.String(20), nullable=False)
    patente_secundaria = db.Column(db.String(20), nullable=True)
    bitren = db.Column(db.String(20), nullable=True)
    tipo_unidad = db.Column(db.String(100), nullable=False)
    dni = db.Column(db.String(20), nullable=False)
    nombre = db.Column(db.String(120), nullable=False)
    transporte = db.Column(db.String(120), nullable=False)
    base_descripcion = db.Column(db.String(180), nullable=True)
    motivo = db.Column(db.Text, nullable=False)
    evidence_original_name = db.Column(db.String(255), nullable=True)
    evidence_stored_name = db.Column(db.String(255), nullable=True)

    ip_origen = db.Column(db.String(100), nullable=True)
    equipo = db.Column(db.String(255), nullable=True)
    usuario_sistema = db.Column(db.String(120), nullable=True)


class AppUser(db.Model):
    __tablename__ = "app_users"

    id = db.Column(Integer, primary_key=True)
    username = db.Column(String(120), unique=True, nullable=False)
    role = db.Column(String(20), nullable=False)
    key_hash = db.Column(String(255), nullable=False)
    active = db.Column(Boolean, nullable=False, default=True)
    must_change_password = db.Column(Boolean, nullable=False, default=False)
    last_password_change_at = db.Column(db.DateTime(timezone=True), nullable=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False)
    updated_at = db.Column(db.DateTime(timezone=True), nullable=False)


class AppSetting(db.Model):
    __tablename__ = "app_settings"

    key = db.Column(String(80), primary_key=True)
    value = db.Column(String(200), nullable=False)
    updated_at = db.Column(db.DateTime(timezone=True), nullable=False)


class UserAuditLog(db.Model):
    __tablename__ = "user_audit_logs"

    id = db.Column(Integer, primary_key=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False)
    actor_username = db.Column(String(120), nullable=False)
    action = db.Column(String(80), nullable=False)
    target_username = db.Column(String(120), nullable=True)
    details = db.Column(Text, nullable=True)


def now_gmt_minus_3() -> datetime:
    return datetime.now(TZ_GMT_MINUS_3)


def safe_text(value: object | None, upper: bool = False) -> str:
    text_value = str(value or "").strip()
    return text_value.upper() if upper else text_value


def allowed_evidence_extension(filename: str) -> bool:
    _, extension = os.path.splitext(filename or "")
    return extension.lower() in ALLOWED_EVIDENCE_EXTENSIONS


def save_evidence_file(file_storage):
    if not file_storage or not file_storage.filename:
        return None, None, None

    original_name = secure_filename(file_storage.filename)
    if not original_name:
        return None, None, "Nombre de archivo invalido"

    if not allowed_evidence_extension(original_name):
        allowed_text = ", ".join(sorted(ALLOWED_EVIDENCE_EXTENSIONS))
        return None, None, f"Extension no permitida. Usa: {allowed_text}"

    file_storage.stream.seek(0, os.SEEK_END)
    size = file_storage.stream.tell()
    file_storage.stream.seek(0)
    if size > MAX_EVIDENCE_BYTES:
        return None, None, "El archivo supera el maximo permitido de 10MB"

    stamp = now_gmt_minus_3().strftime("%Y%m%d_%H%M%S_%f")
    unique_name = f"{stamp}_{original_name}"
    stored_name = secure_filename(unique_name)
    file_path = os.path.join(EVIDENCE_DIR, stored_name)
    file_storage.save(file_path)
    return original_name, stored_name, None


def normalize_plate(value: str) -> str:
    return "".join(ch for ch in safe_text(value, upper=True) if ch.isalnum())


def prettify_unit_type(raw_value: str) -> str:
    value = safe_text(raw_value, upper=False)
    if not value:
        return ""
    cleaned = value.replace("_", " ").replace("-", " ")
    return " ".join(cleaned.split())


def find_header_index(headers: list, expected_name: str) -> int | None:
    expected = safe_text(expected_name).lower()
    for index, name in enumerate(headers):
        if safe_text(name).lower() == expected:
            return index
    return None


def _read_reference_data_from_file(file_path: str) -> tuple[list[str], dict[str, str], dict[str, str]]:
    wb = load_workbook(file_path, data_only=True, read_only=True)

    try:
        plantas = wb["Plantas"] if "Plantas" in wb.sheetnames else None
        transportes = wb["Transportes"] if "Transportes" in wb.sheetnames else None

        bases = []
        transport_map = {}
        unit_type_map = {}

        if plantas:
            headers = list(next(plantas.iter_rows(min_row=1, max_row=1, values_only=True), ()))
            base_idx = find_header_index(headers, "Descripción de Base")
            if base_idx is not None:
                bases_set = set()
                for row in plantas.iter_rows(min_row=2, values_only=True):
                    if not row or base_idx >= len(row):
                        continue
                    base_name = safe_text(str(row[base_idx] or ""))
                    if base_name:
                        bases_set.add(base_name)
                bases = sorted(bases_set)

        if transportes:
            headers = list(next(transportes.iter_rows(min_row=1, max_row=1, values_only=True), ()))
            plate_idx = find_header_index(headers, "Inicial de equipo")
            trans_idx = find_header_index(headers, "Transportista")
            type_idx = find_header_index(headers, "ID de tipo de equipo")

            if plate_idx is not None and (trans_idx is not None or type_idx is not None):
                for row in transportes.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    plate_raw = row[plate_idx] if plate_idx < len(row) else None

                    normalized_plate = normalize_plate(str(plate_raw or ""))
                    if not normalized_plate:
                        continue

                    if trans_idx is not None:
                        transport_raw = row[trans_idx] if trans_idx < len(row) else None
                        transport_name = safe_text(str(transport_raw or ""))
                        if transport_name:
                            existing = transport_map.get(normalized_plate)
                            if not existing:
                                transport_map[normalized_plate] = transport_name
                            elif safe_text(existing).upper() == "NOTFOUND" and safe_text(transport_name).upper() != "NOTFOUND":
                                transport_map[normalized_plate] = transport_name

                    if type_idx is not None:
                        unit_raw = row[type_idx] if type_idx < len(row) else None
                        unit_type = prettify_unit_type(str(unit_raw or ""))
                        if unit_type:
                            if not unit_type_map.get(normalized_plate):
                                unit_type_map[normalized_plate] = unit_type

        return bases, transport_map, unit_type_map
    finally:
        wb.close()


def load_reference_data() -> tuple[list[str], dict[str, str], dict[str, str]]:
    if not os.path.exists(REFERENCE_DATA_FILE):
        return [], {}, {}

    try:
        mtime = os.path.getmtime(REFERENCE_DATA_FILE)
    except OSError:
        return REFERENCE_CACHE["bases"], REFERENCE_CACHE["transport_by_plate"], REFERENCE_CACHE["unit_type_by_plate"]

    if REFERENCE_CACHE["mtime"] == mtime:
        return REFERENCE_CACHE["bases"], REFERENCE_CACHE["transport_by_plate"], REFERENCE_CACHE["unit_type_by_plate"]

    try:
        bases, transport_map, unit_type_map = _read_reference_data_from_file(REFERENCE_DATA_FILE)
    except PermissionError:
        temp_path = None
        try:
            fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="data_ref_")
            os.close(fd)

            # En Windows, Excel puede bloquear lectura directa desde Python.
            # Copy-Item suele poder duplicar el archivo con mejor tolerancia de lock.
            copy_result = subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    f"Copy-Item -Path '{REFERENCE_DATA_FILE}' -Destination '{temp_path}' -Force",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            if copy_result.returncode != 0:
                return REFERENCE_CACHE["bases"], REFERENCE_CACHE["transport_by_plate"], REFERENCE_CACHE["unit_type_by_plate"]

            bases, transport_map, unit_type_map = _read_reference_data_from_file(temp_path)
        except OSError:
            return REFERENCE_CACHE["bases"], REFERENCE_CACHE["transport_by_plate"], REFERENCE_CACHE["unit_type_by_plate"]
        finally:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
    except OSError:
        return REFERENCE_CACHE["bases"], REFERENCE_CACHE["transport_by_plate"], REFERENCE_CACHE["unit_type_by_plate"]

    REFERENCE_CACHE["mtime"] = mtime
    REFERENCE_CACHE["bases"] = bases
    REFERENCE_CACHE["transport_by_plate"] = transport_map
    REFERENCE_CACHE["unit_type_by_plate"] = unit_type_map
    return bases, transport_map, unit_type_map


def lookup_transport_from_reference(*plates: str) -> tuple[str | None, str | None]:
    _, transport_map, _ = load_reference_data()
    for plate in plates:
        normalized = normalize_plate(plate)
        if not normalized:
            continue
        transport = transport_map.get(normalized)
        if transport:
            return transport, normalized
    return None, None


def lookup_unit_type_from_reference(primary: str, secondary: str, bitren: str) -> tuple[str | None, dict]:
    _, _, type_map = load_reference_data()

    ordered = [
        normalize_plate(primary),
        normalize_plate(secondary),
        normalize_plate(bitren),
    ]

    types = []
    detail: dict[str, str | None] = {
        "patente_primaria": None,
        "patente_secundaria": None,
        "bitren": None,
    }

    keys = ["patente_primaria", "patente_secundaria", "bitren"]
    for key, plate in zip(keys, ordered):
        if not plate:
            continue
        unit_type = type_map.get(plate)
        if not unit_type:
            continue
        detail[key] = unit_type
        types.append(unit_type)

    if not types:
        return None, detail

    return " / ".join(types), detail


def is_legacy_sha256_hash(value: str) -> bool:
    text_value = (value or "").strip()
    return len(text_value) == 64 and all(ch in "0123456789abcdefABCDEF" for ch in text_value)


def apply_pepper(raw_key: str) -> str:
    value = raw_key or ""
    if not AUTH_PEPPER:
        return value
    return f"{value}{AUTH_PEPPER}"


def hash_key(raw_key: str) -> str:
    return password_hasher.hash(apply_pepper(raw_key))


def verify_key_hash(raw_key: str, stored_hash: str) -> bool:
    value = (stored_hash or "").strip()

    if value.startswith("$argon2id$"):
        try:
            return bool(password_hasher.verify(value, apply_pepper(raw_key)))
        except (VerifyMismatchError, InvalidHash, ValueError):
            return False

    return False


def get_hash_scheme(stored_hash: str) -> str:
    value = (stored_hash or "").strip()
    if value.startswith("$argon2id$"):
        return ARGON2_SCHEME
    if is_legacy_sha256_hash(value):
        return "legacy_sha256"
    return "unknown"


def parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "si", "yes", "on"}


def parse_int(value: str | None, default: int) -> int:
    try:
        return int(str(value))
    except (TypeError, ValueError):
        return default


def get_setting_value(key: str, default: str) -> str:
    setting = AppSetting.query.filter_by(key=key).first()
    return setting.value if setting else default


def set_setting_value(key: str, value: str):
    setting = AppSetting.query.filter_by(key=key).first()
    now = now_gmt_minus_3()
    if setting:
        setting.value = value
        setting.updated_at = now
    else:
        new_setting = AppSetting()
        new_setting.key = key
        new_setting.value = value
        new_setting.updated_at = now
        db.session.add(new_setting)


def get_role_permissions_matrix() -> dict:
    raw = get_setting_value("role_permissions_json", "")
    matrix = clone_default_permissions()

    if raw:
        try:
            payload = json.loads(raw)
            if isinstance(payload, dict):
                for role, permissions in payload.items():
                    role_name = safe_text(role, upper=True)
                    if role_name not in ROLE_TYPES or not isinstance(permissions, dict):
                        continue
                    for permission in PERMISSION_KEYS:
                        if permission in permissions:
                            matrix[role_name][permission] = bool(permissions[permission])
        except json.JSONDecodeError:
            pass

    return matrix


def save_role_permissions_matrix(matrix: dict):
    normalized = {}
    for role in sorted(ROLE_TYPES):
        role_permissions = matrix.get(role, {}) if isinstance(matrix.get(role), dict) else {}
        normalized[role] = {
            permission: bool(role_permissions.get(permission, DEFAULT_ROLE_PERMISSIONS[role][permission]))
            for permission in sorted(PERMISSION_KEYS)
        }

    # Evita lockout administrativo.
    normalized["ADMIN"]["can_manage_permissions"] = True
    normalized["ADMIN"]["can_access_admin_panel"] = True

    set_setting_value("role_permissions_json", json.dumps(normalized, ensure_ascii=True, sort_keys=True))


def get_permissions_for_role(role: str) -> dict:
    role_name = safe_text(role, upper=True)
    matrix = get_role_permissions_matrix()
    return matrix.get(role_name, {permission: False for permission in PERMISSION_KEYS})


def has_permission(role: str, permission: str) -> bool:
    if permission not in PERMISSION_KEYS:
        return False
    role_permissions = get_permissions_for_role(role)
    return bool(role_permissions.get(permission, False))


def get_unlock_allowed_roles() -> list[str]:
    return sorted(role for role in ROLE_TYPES if has_permission(role, "can_unlock_requests"))


def ensure_setting_default(key: str, value: str):
    setting = AppSetting.query.filter_by(key=key).first()
    if setting:
        return
    new_setting = AppSetting()
    new_setting.key = key
    new_setting.value = value
    new_setting.updated_at = now_gmt_minus_3()
    db.session.add(new_setting)


def get_password_policy() -> dict:
    return {
        "min_length": max(6, parse_int(get_setting_value("password_min_length", str(DEFAULT_PASSWORD_POLICY["min_length"])), DEFAULT_PASSWORD_POLICY["min_length"])),
        "require_uppercase": parse_bool(get_setting_value("password_require_uppercase", "1" if DEFAULT_PASSWORD_POLICY["require_uppercase"] else "0"), DEFAULT_PASSWORD_POLICY["require_uppercase"]),
        "require_lowercase": parse_bool(get_setting_value("password_require_lowercase", "1" if DEFAULT_PASSWORD_POLICY["require_lowercase"] else "0"), DEFAULT_PASSWORD_POLICY["require_lowercase"]),
        "require_digit": parse_bool(get_setting_value("password_require_digit", "1" if DEFAULT_PASSWORD_POLICY["require_digit"] else "0"), DEFAULT_PASSWORD_POLICY["require_digit"]),
        "require_symbol": parse_bool(get_setting_value("password_require_symbol", "1" if DEFAULT_PASSWORD_POLICY["require_symbol"] else "0"), DEFAULT_PASSWORD_POLICY["require_symbol"]),
        "expires_days": max(0, parse_int(get_setting_value("password_expires_days", str(DEFAULT_PASSWORD_POLICY["expires_days"])), DEFAULT_PASSWORD_POLICY["expires_days"])),
    }


def split_email_list(raw_value: str) -> list[str]:
    text_value = safe_text(raw_value)
    if not text_value:
        return []

    normalized = text_value.replace(";", ",")
    values = []
    for item in normalized.split(","):
        email = safe_text(item).lower()
        if email and email not in values:
            values.append(email)
    return values


def normalize_email(value: object) -> str:
    text_value = safe_text(value).lower()
    if not text_value:
        return ""

    _, parsed = parseaddr(text_value)
    email_value = safe_text(parsed).lower()
    if not email_value or "@" not in email_value:
        return ""

    local_part, _, domain_part = email_value.partition("@")
    if not local_part or not domain_part or "." not in domain_part:
        return ""

    return email_value


def parse_smtp_port(value: object, default: int = 587) -> int:
    try:
        port = int(str(value))
    except (TypeError, ValueError):
        return default
    if 1 <= port <= 65535:
        return port
    return default


def get_notification_settings(include_secret: bool = False) -> dict:
    smtp_password = get_setting_value("mail_smtp_password", DEFAULT_NOTIFICATION_SETTINGS["smtp_password"])
    settings = {
        "enabled": parse_bool(get_setting_value("mail_enabled", "0"), DEFAULT_NOTIFICATION_SETTINGS["enabled"]),
        "smtp_host": get_setting_value("mail_smtp_host", DEFAULT_NOTIFICATION_SETTINGS["smtp_host"]),
        "smtp_port": parse_smtp_port(get_setting_value("mail_smtp_port", str(DEFAULT_NOTIFICATION_SETTINGS["smtp_port"])), DEFAULT_NOTIFICATION_SETTINGS["smtp_port"]),
        "smtp_username": get_setting_value("mail_smtp_username", DEFAULT_NOTIFICATION_SETTINGS["smtp_username"]),
        "smtp_password": smtp_password if include_secret else "",
        "smtp_password_set": bool(smtp_password),
        "smtp_use_tls": parse_bool(get_setting_value("mail_smtp_use_tls", "1"), DEFAULT_NOTIFICATION_SETTINGS["smtp_use_tls"]),
        "smtp_use_ssl": parse_bool(get_setting_value("mail_smtp_use_ssl", "0"), DEFAULT_NOTIFICATION_SETTINGS["smtp_use_ssl"]),
        "from_name": get_setting_value("mail_from_name", DEFAULT_NOTIFICATION_SETTINGS["from_name"]),
        "from_email": get_setting_value("mail_from_email", DEFAULT_NOTIFICATION_SETTINGS["from_email"]),
        "to_addresses": get_setting_value("mail_to_addresses", DEFAULT_NOTIFICATION_SETTINGS["to_addresses"]),
        "cc_addresses": get_setting_value("mail_cc_addresses", DEFAULT_NOTIFICATION_SETTINGS["cc_addresses"]),
        "subject_bloquear": get_setting_value("mail_subject_bloquear", DEFAULT_NOTIFICATION_SETTINGS["subject_bloquear"]),
        "subject_desbloquear": get_setting_value("mail_subject_desbloquear", DEFAULT_NOTIFICATION_SETTINGS["subject_desbloquear"]),
        "subject_novedad": get_setting_value("mail_subject_novedad", DEFAULT_NOTIFICATION_SETTINGS["subject_novedad"]),
        "subject_incidencia": get_setting_value("mail_subject_incidencia", DEFAULT_NOTIFICATION_SETTINGS["subject_incidencia"]),
    }
    return settings


def save_notification_settings(payload: dict):
    current = get_notification_settings(include_secret=True)

    enabled = bool(payload.get("enabled", current["enabled"]))
    smtp_host = safe_text(payload.get("smtp_host", current["smtp_host"]))
    smtp_port = parse_smtp_port(payload.get("smtp_port", current["smtp_port"]), current["smtp_port"])
    smtp_username = safe_text(payload.get("smtp_username", current["smtp_username"]))
    smtp_password_new = payload.get("smtp_password", None)
    smtp_password = current["smtp_password"] if smtp_password_new in (None, "") else str(smtp_password_new)
    smtp_use_tls = bool(payload.get("smtp_use_tls", current["smtp_use_tls"]))
    smtp_use_ssl = bool(payload.get("smtp_use_ssl", current["smtp_use_ssl"]))
    from_name = safe_text(payload.get("from_name", current["from_name"]))
    from_email = safe_text(payload.get("from_email", current["from_email"])).lower()
    to_addresses = safe_text(payload.get("to_addresses", current["to_addresses"]))
    cc_addresses = safe_text(payload.get("cc_addresses", current["cc_addresses"]))
    subject_bloquear = safe_text(payload.get("subject_bloquear", current["subject_bloquear"]))
    subject_desbloquear = safe_text(payload.get("subject_desbloquear", current["subject_desbloquear"]))
    subject_novedad = safe_text(payload.get("subject_novedad", current["subject_novedad"]))
    subject_incidencia = safe_text(payload.get("subject_incidencia", current["subject_incidencia"]))

    if enabled:
        if not smtp_host:
            return False, "SMTP host es obligatorio cuando mail esta habilitado"
        if not split_email_list(to_addresses):
            return False, "Debes cargar al menos un destinatario TO cuando mail esta habilitado"
        if smtp_use_ssl and smtp_use_tls:
            return False, "No puedes activar SSL y TLS al mismo tiempo"

    set_setting_value("mail_enabled", "1" if enabled else "0")
    set_setting_value("mail_smtp_host", smtp_host)
    set_setting_value("mail_smtp_port", str(smtp_port))
    set_setting_value("mail_smtp_username", smtp_username)
    set_setting_value("mail_smtp_password", smtp_password)
    set_setting_value("mail_smtp_use_tls", "1" if smtp_use_tls else "0")
    set_setting_value("mail_smtp_use_ssl", "1" if smtp_use_ssl else "0")
    set_setting_value("mail_from_name", from_name)
    set_setting_value("mail_from_email", from_email)
    set_setting_value("mail_to_addresses", to_addresses)
    set_setting_value("mail_cc_addresses", cc_addresses)
    set_setting_value("mail_subject_bloquear", subject_bloquear or DEFAULT_NOTIFICATION_SETTINGS["subject_bloquear"])
    set_setting_value("mail_subject_desbloquear", subject_desbloquear or DEFAULT_NOTIFICATION_SETTINGS["subject_desbloquear"])
    set_setting_value("mail_subject_novedad", subject_novedad or DEFAULT_NOTIFICATION_SETTINGS["subject_novedad"])
    set_setting_value("mail_subject_incidencia", subject_incidencia or DEFAULT_NOTIFICATION_SETTINGS["subject_incidencia"])

    return True, None


def get_notification_subject(action_type: str, settings: dict) -> str:
    action = safe_text(action_type, upper=True)
    mapping = {
        "BLOQUEAR": settings.get("subject_bloquear", DEFAULT_NOTIFICATION_SETTINGS["subject_bloquear"]),
        "DESBLOQUEAR": settings.get("subject_desbloquear", DEFAULT_NOTIFICATION_SETTINGS["subject_desbloquear"]),
        "NOVEDAD": settings.get("subject_novedad", DEFAULT_NOTIFICATION_SETTINGS["subject_novedad"]),
        "INCIDENCIA": settings.get("subject_incidencia", DEFAULT_NOTIFICATION_SETTINGS["subject_incidencia"]),
    }
    return mapping.get(action, f"[Registro] Nuevo evento {action}")


def build_notification_body(item: BlockRequest, requester_email: str | None = None) -> str:
    lines = [
        "Se registro un nuevo evento en el sistema.",
        "",
        f"ID: {item.id}",
        f"Accion: {item.action_type}",
        f"Fecha sistema (GMT-3): {item.created_at_system.isoformat()}",
        f"Fecha evento (GMT-3): {item.event_datetime.isoformat()}",
        f"Patente primaria: {item.patente_primaria}",
        f"Patente secundaria: {item.patente_secundaria or '-'}",
        f"Bitren: {item.bitren or '-'}",
        f"Tipo unidad: {item.tipo_unidad}",
        f"DNI: {item.dni}",
        f"Nombre: {item.nombre}",
        f"Transporte: {item.transporte}",
        f"Base: {item.base_descripcion or '-'}",
        f"Nro carga: {item.carga_nro or '-'}",
        f"Autorizado por: {item.authorized_by_user or '-'}",
        f"Email solicitante: {requester_email or '-'}",
        f"Usuario sistema: {item.usuario_sistema or '-'}",
        f"IP origen: {item.ip_origen or '-'}",
        "",
        "Motivo:",
        item.motivo or "-",
    ]
    return "\n".join(lines)


def send_notification_email(item: BlockRequest, requester_email: str | None = None) -> tuple[bool, str | None]:
    settings = get_notification_settings(include_secret=True)
    if not settings["enabled"]:
        return True, None

    to_addresses = split_email_list(settings["to_addresses"])
    cc_addresses = split_email_list(settings["cc_addresses"])
    requester_copy = normalize_email(requester_email)
    if requester_copy and requester_copy not in to_addresses and requester_copy not in cc_addresses:
        cc_addresses.append(requester_copy)
    if not to_addresses:
        return False, "No hay destinatarios TO configurados"

    smtp_host = safe_text(settings["smtp_host"])
    if not smtp_host:
        return False, "SMTP host no configurado"

    sender_email = safe_text(settings["from_email"]).lower() or safe_text(settings["smtp_username"]).lower()
    if not sender_email:
        return False, "No hay from_email ni smtp_username configurado"

    message = EmailMessage()
    message["Subject"] = get_notification_subject(item.action_type, settings)
    message["From"] = f"{safe_text(settings['from_name'])} <{sender_email}>" if safe_text(settings["from_name"]) else sender_email
    message["To"] = ", ".join(to_addresses)
    if cc_addresses:
        message["Cc"] = ", ".join(cc_addresses)
    message.set_content(build_notification_body(item, requester_copy or None), subtype="plain", charset="utf-8")

    recipients = to_addresses + cc_addresses
    smtp_port = parse_smtp_port(settings["smtp_port"], 587)
    smtp_username = safe_text(settings["smtp_username"])
    smtp_password = str(settings["smtp_password"] or "")
    use_tls = bool(settings["smtp_use_tls"])
    use_ssl = bool(settings["smtp_use_ssl"])

    try:
        if use_ssl:
            with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=20) as server:
                if smtp_username:
                    server.login(smtp_username, smtp_password)
                server.send_message(message, from_addr=sender_email, to_addrs=recipients)
        else:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
                server.ehlo()
                if use_tls:
                    server.starttls()
                    server.ehlo()
                if smtp_username:
                    server.login(smtp_username, smtp_password)
                server.send_message(message, from_addr=sender_email, to_addrs=recipients)
    except Exception as exc:
        return False, str(exc)

    return True, None


def validate_password_strength(password: str, policy: dict) -> str | None:
    if len(password) < policy["min_length"]:
        return f"La clave debe tener al menos {policy['min_length']} caracteres"
    if policy["require_uppercase"] and not any(ch.isupper() for ch in password):
        return "La clave debe incluir al menos una mayuscula"
    if policy["require_lowercase"] and not any(ch.islower() for ch in password):
        return "La clave debe incluir al menos una minuscula"
    if policy["require_digit"] and not any(ch.isdigit() for ch in password):
        return "La clave debe incluir al menos un numero"
    if policy["require_symbol"] and not any(not ch.isalnum() for ch in password):
        return "La clave debe incluir al menos un simbolo"
    return None


def ensure_tz(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=TZ_GMT_MINUS_3)
    return value.astimezone(TZ_GMT_MINUS_3)


def get_user_password_issue(user: AppUser) -> str | None:
    if user.must_change_password:
        return "must_change_password"

    policy = get_password_policy()
    expires_days = policy["expires_days"]
    if expires_days <= 0:
        return None

    last_change = ensure_tz(user.last_password_change_at)
    if not last_change:
        return "password_expired"

    age_days = (now_gmt_minus_3() - last_change).days
    if age_days >= expires_days:
        return "password_expired"
    return None


def get_days_until_expiry(user: AppUser) -> int | None:
    policy = get_password_policy()
    expires_days = policy["expires_days"]
    if expires_days <= 0:
        return None

    if user.must_change_password:
        return -1

    last_change = ensure_tz(user.last_password_change_at)
    if not last_change:
        return -1

    age_days = (now_gmt_minus_3() - last_change).days
    return expires_days - age_days


def get_admin_hash() -> str | None:
    if not os.path.exists(ADMIN_HASH_FILE) and os.path.exists(LEGACY_ADMIN_HASH_FILE):
        try:
            os.makedirs(os.path.dirname(ADMIN_HASH_FILE), exist_ok=True)
            os.replace(LEGACY_ADMIN_HASH_FILE, ADMIN_HASH_FILE)
        except OSError:
            pass

    if not os.path.exists(ADMIN_HASH_FILE):
        return None

    try:
        with open(ADMIN_HASH_FILE, "r", encoding="utf-8") as file:
            content = file.read().strip()
    except OSError:
        return None

    if not content:
        return None

    if content.startswith("$argon2id$"):
        return content

    return None


def verify_admin_key(raw_key: str) -> bool:
    expected = get_admin_hash()
    if not expected:
        return False
    return verify_key_hash(safe_text(raw_key), expected)


def can_use_legacy_admin(username: str) -> bool:
    normalized = safe_text(username, upper=True)
    return normalized in {"ADMIN", "LEGACY_ADMIN"}


def has_user_auth() -> bool:
    return db.session.query(AppUser.id).first() is not None


def is_admin_enabled() -> bool:
    return has_user_auth() or get_admin_hash() is not None


def set_session_auth(username: str, role: str, source: str):
    session[SESSION_USERNAME_KEY] = safe_text(username)
    session[SESSION_ROLE_KEY] = safe_text(role, upper=True)
    session[SESSION_SOURCE_KEY] = safe_text(source)


def clear_session_auth():
    session.pop(SESSION_USERNAME_KEY, None)
    session.pop(SESSION_ROLE_KEY, None)
    session.pop(SESSION_SOURCE_KEY, None)


def get_session_principal() -> dict | None:
    username = safe_text(session.get(SESSION_USERNAME_KEY, ""))
    role = safe_text(session.get(SESSION_ROLE_KEY, ""), upper=True)
    source = safe_text(session.get(SESSION_SOURCE_KEY, ""))

    if not username or role not in ROLE_TYPES:
        return None

    permissions = get_permissions_for_role(role)
    if not permissions.get("can_access_admin_panel", False):
        return None

    return {
        "username": username,
        "role": role,
        "source": source,
        "permissions": permissions,
    }


def authenticate_operation_session():
    principal = get_session_principal()
    if not principal:
        return None, jsonify({"error": "Sesion requerida. Inicia login para continuar."}), 401
    return principal, None, 200


def verify_user_credentials(username: str, raw_key: str) -> AppUser | None:
    if LEGACY_ONLY_ADMIN:
        return None

    normalized = safe_text(username)
    if not normalized:
        return None

    user = AppUser.query.filter(func.lower(AppUser.username) == normalized.lower()).first()
    if not user or not user.active:
        return None

    normalized_key = safe_text(raw_key)
    # Evita que una clave de usuario normal coincida con la admin key legacy.
    if not can_use_legacy_admin(normalized) and verify_admin_key(normalized_key):
        return None

    if verify_key_hash(normalized_key, user.key_hash):
        if get_hash_scheme(user.key_hash) != ARGON2_SCHEME:
            user.key_hash = hash_key(normalized_key)
            user.updated_at = now_gmt_minus_3()
            db.session.commit()
        return user

    return None


def authenticate_admin_request(required_roles: set[str] | None = None, required_permission: str | None = None):
    admin_user = safe_text(
        request.headers.get("X-Admin-User", "") or request.args.get("admin_user", "")
    )
    admin_key = safe_text(
        request.headers.get("X-Admin-Key", "") or request.args.get("admin_key", "")
    )

    principal = None

    if not LEGACY_ONLY_ADMIN:
        user = verify_user_credentials(admin_user, admin_key)
        if user:
            password_issue = get_user_password_issue(user)
            permissions = get_permissions_for_role(user.role)
            principal = {
                "username": user.username,
                "role": user.role,
                "source": "user",
                "must_change_password": password_issue is not None,
                "password_issue": password_issue,
                "permissions": permissions,
            }

    if not principal and can_use_legacy_admin(admin_user) and verify_admin_key(admin_key):
        permissions = get_permissions_for_role("ADMIN")
        principal = {
            "username": "legacy_admin",
            "role": "ADMIN",
            "source": "legacy",
            "must_change_password": False,
            "permissions": permissions,
        }

    if not principal:
        return None, jsonify({"error": "Credenciales admin invalidas"}), 403

    if principal["source"] == "user" and principal.get("must_change_password"):
        return None, jsonify({"error": "Debe cambiar su clave antes de continuar"}), 403

    if required_roles and principal["role"] not in required_roles:
        return None, jsonify({"error": "Sin permisos para esta accion"}), 403

    if required_permission and not bool(principal.get("permissions", {}).get(required_permission, False)):
        return None, jsonify({"error": "Sin permisos para esta accion"}), 403

    return principal, None, 200


def parse_event_datetime(value: object) -> datetime:
    date_value = datetime.fromisoformat(str(value))
    if date_value.tzinfo is None:
        date_value = date_value.replace(tzinfo=TZ_GMT_MINUS_3)
    return date_value.astimezone(TZ_GMT_MINUS_3)


def serialize_request(item: BlockRequest) -> dict:
    return {
        "id": item.id,
        "created_at_system": item.created_at_system.isoformat(),
        "event_datetime": item.event_datetime.isoformat(),
        "action_type": item.action_type,
        "related_request_id": item.related_request_id,
        "requested_by_user": item.usuario_sistema,
        "authorized_by_user": item.authorized_by_user,
        "carga_nro": item.carga_nro,
        "patente_primaria": item.patente_primaria,
        "patente_secundaria": item.patente_secundaria,
        "bitren": item.bitren,
        "tipo_unidad": item.tipo_unidad,
        "dni": item.dni,
        "nombre": item.nombre,
        "transporte": item.transporte,
        "base_descripcion": item.base_descripcion,
        "motivo": item.motivo,
        "evidence_original_name": item.evidence_original_name,
        "has_evidence": bool(item.evidence_stored_name),
        "ip_origen": item.ip_origen,
        "equipo": item.equipo,
        "usuario_sistema": item.usuario_sistema,
    }


def serialize_user(user: AppUser) -> dict:
    return {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "active": user.active,
        "hash_scheme": get_hash_scheme(user.key_hash),
        "must_change_password": user.must_change_password,
        "password_issue": get_user_password_issue(user),
        "last_password_change_at": user.last_password_change_at.isoformat() if user.last_password_change_at else None,
        "created_at": user.created_at.isoformat(),
        "updated_at": user.updated_at.isoformat(),
    }


def serialize_password_health_user(user: AppUser) -> dict:
    return {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "active": user.active,
        "hash_scheme": get_hash_scheme(user.key_hash),
        "must_change_password": user.must_change_password,
        "password_issue": get_user_password_issue(user),
        "days_until_expiry": get_days_until_expiry(user),
        "last_password_change_at": user.last_password_change_at.isoformat() if user.last_password_change_at else None,
    }


def serialize_audit(item: UserAuditLog) -> dict:
    return {
        "id": item.id,
        "created_at": item.created_at.isoformat(),
        "actor_username": item.actor_username,
        "action": item.action,
        "target_username": item.target_username,
        "details": item.details,
    }


def add_audit(actor_username: str, action: str, target_username: str | None = None, details: str | None = None):
    audit = UserAuditLog()
    audit.created_at = now_gmt_minus_3()
    audit.actor_username = safe_text(actor_username) or "sistema"
    audit.action = safe_text(action, upper=True)
    audit.target_username = safe_text(target_username) if target_username else None
    audit.details = safe_text(details) if details else None
    db.session.add(audit)


def ensure_block_request_schema():
    inspector = inspect(db.engine)
    existing_columns = {col["name"] for col in inspector.get_columns("block_requests")}

    column_statements = {
        "patente_secundaria": "ALTER TABLE block_requests ADD COLUMN patente_secundaria VARCHAR(20)",
        "bitren": "ALTER TABLE block_requests ADD COLUMN bitren VARCHAR(20)",
        "action_type": "ALTER TABLE block_requests ADD COLUMN action_type VARCHAR(20)",
        "related_request_id": "ALTER TABLE block_requests ADD COLUMN related_request_id INTEGER",
        "authorized_by_user": "ALTER TABLE block_requests ADD COLUMN authorized_by_user VARCHAR(120)",
        "base_descripcion": "ALTER TABLE block_requests ADD COLUMN base_descripcion VARCHAR(180)",
        "evidence_original_name": "ALTER TABLE block_requests ADD COLUMN evidence_original_name VARCHAR(255)",
        "evidence_stored_name": "ALTER TABLE block_requests ADD COLUMN evidence_stored_name VARCHAR(255)",
    }

    for column_name, ddl in column_statements.items():
        if column_name not in existing_columns:
            db.session.execute(text(ddl))

    db.session.execute(text("UPDATE block_requests SET action_type='BLOQUEAR' WHERE action_type IS NULL"))
    db.session.commit()


def ensure_user_schema():
    inspector = inspect(db.engine)
    existing_columns = {col["name"] for col in inspector.get_columns("app_users")}

    if "must_change_password" not in existing_columns:
        db.session.execute(text("ALTER TABLE app_users ADD COLUMN must_change_password BOOLEAN"))
        db.session.execute(text("UPDATE app_users SET must_change_password=0 WHERE must_change_password IS NULL"))
        db.session.commit()

    if "last_password_change_at" not in existing_columns:
        db.session.execute(text("ALTER TABLE app_users ADD COLUMN last_password_change_at DATETIME"))
        db.session.commit()


def ensure_settings_defaults():
    ensure_setting_default("password_min_length", str(DEFAULT_PASSWORD_POLICY["min_length"]))
    ensure_setting_default("password_require_uppercase", "1" if DEFAULT_PASSWORD_POLICY["require_uppercase"] else "0")
    ensure_setting_default("password_require_lowercase", "1" if DEFAULT_PASSWORD_POLICY["require_lowercase"] else "0")
    ensure_setting_default("password_require_digit", "1" if DEFAULT_PASSWORD_POLICY["require_digit"] else "0")
    ensure_setting_default("password_require_symbol", "1" if DEFAULT_PASSWORD_POLICY["require_symbol"] else "0")
    ensure_setting_default("password_expires_days", str(DEFAULT_PASSWORD_POLICY["expires_days"]))
    ensure_setting_default("role_permissions_json", json.dumps(clone_default_permissions(), ensure_ascii=True, sort_keys=True))
    ensure_setting_default("mail_enabled", "1" if DEFAULT_NOTIFICATION_SETTINGS["enabled"] else "0")
    ensure_setting_default("mail_smtp_host", DEFAULT_NOTIFICATION_SETTINGS["smtp_host"])
    ensure_setting_default("mail_smtp_port", str(DEFAULT_NOTIFICATION_SETTINGS["smtp_port"]))
    ensure_setting_default("mail_smtp_username", DEFAULT_NOTIFICATION_SETTINGS["smtp_username"])
    ensure_setting_default("mail_smtp_password", DEFAULT_NOTIFICATION_SETTINGS["smtp_password"])
    ensure_setting_default("mail_smtp_use_tls", "1" if DEFAULT_NOTIFICATION_SETTINGS["smtp_use_tls"] else "0")
    ensure_setting_default("mail_smtp_use_ssl", "1" if DEFAULT_NOTIFICATION_SETTINGS["smtp_use_ssl"] else "0")
    ensure_setting_default("mail_from_name", DEFAULT_NOTIFICATION_SETTINGS["from_name"])
    ensure_setting_default("mail_from_email", DEFAULT_NOTIFICATION_SETTINGS["from_email"])
    ensure_setting_default("mail_to_addresses", DEFAULT_NOTIFICATION_SETTINGS["to_addresses"])
    ensure_setting_default("mail_cc_addresses", DEFAULT_NOTIFICATION_SETTINGS["cc_addresses"])
    ensure_setting_default("mail_subject_bloquear", DEFAULT_NOTIFICATION_SETTINGS["subject_bloquear"])
    ensure_setting_default("mail_subject_desbloquear", DEFAULT_NOTIFICATION_SETTINGS["subject_desbloquear"])
    ensure_setting_default("mail_subject_novedad", DEFAULT_NOTIFICATION_SETTINGS["subject_novedad"])
    ensure_setting_default("mail_subject_incidencia", DEFAULT_NOTIFICATION_SETTINGS["subject_incidencia"])
    db.session.commit()


@app.route("/")
def index():
    return render_template("login_landing.html")


@app.route("/operacion")
def operation_page():
    if not get_session_principal():
        return redirect(url_for("index"))
    return render_template("index.html")


@app.route("/admin")
def admin_page():
    return render_template("admin.html")


@app.route("/api/requests", methods=["POST"])
def create_request():
    principal, error_response, status_code = authenticate_operation_session()
    if error_response:
        return error_response, status_code
    assert principal is not None

    is_multipart = request.content_type and "multipart/form-data" in request.content_type.lower()
    payload = request.form.to_dict() if is_multipart else (request.get_json(silent=True) or {})

    required = [
        "event_datetime",
        "patente_primaria",
        "tipo_unidad",
        "dni",
        "nombre",
        "transporte",
        "motivo",
    ]
    missing = [field for field in required if not safe_text(payload.get(field))]
    if missing:
        return jsonify({"error": "Faltan campos obligatorios", "missing": missing}), 400

    action_type = safe_text(payload.get("action_type"), upper=True) or "BLOQUEAR"
    if action_type not in ACTION_TYPES:
        return jsonify({"error": "Accion invalida. Use BLOQUEAR, DESBLOQUEAR, NOVEDAD o INCIDENCIA"}), 400

    try:
        event_datetime = parse_event_datetime(payload.get("event_datetime"))
    except ValueError:
        return jsonify({"error": "Fecha y hora del evento invalida"}), 400

    auth_user = safe_text(payload.get("auth_user"))
    auth_key = safe_text(payload.get("auth_key"))

    if action_type == "DESBLOQUEAR":
        auth_user_obj = verify_user_credentials(auth_user, auth_key)
        if not auth_user_obj:
            return jsonify({"error": "Usuario/clave invalidos para desbloquear"}), 403
        issue = get_user_password_issue(auth_user_obj)
        if issue:
            return jsonify({"error": "Usuario autorizado con clave vencida o pendiente de cambio"}), 403
        if not has_permission(auth_user_obj.role, "can_unlock_requests"):
            return jsonify({"error": "El usuario autorizado no tiene permiso para desbloquear"}), 403

    related_request_id = payload.get("related_request_id")
    if related_request_id not in (None, ""):
        try:
            related_request_id = int(related_request_id)
        except (TypeError, ValueError):
            return jsonify({"error": "ID de linea asociada invalido"}), 400
    else:
        related_request_id = None

    forwarded_for = request.headers.get("X-Forwarded-For", "")
    ip_origen = forwarded_for.split(",")[0].strip() if forwarded_for else request.remote_addr
    requester_email = safe_text(payload.get("email_usuario")).lower()
    if requester_email and not normalize_email(requester_email):
        return jsonify({"error": "Email solicitante invalido"}), 400

    evidence_original_name = None
    evidence_stored_name = None
    if is_multipart:
        evidence_original_name, evidence_stored_name, evidence_error = save_evidence_file(request.files.get("evidence_file"))
        if evidence_error:
            return jsonify({"error": evidence_error}), 400

    item = BlockRequest()
    item.created_at_system = now_gmt_minus_3()
    item.event_datetime = event_datetime
    item.action_type = action_type
    item.related_request_id = related_request_id
    item.authorized_by_user = auth_user if action_type == "DESBLOQUEAR" else None
    item.carga_nro = safe_text(payload.get("carga_nro"))
    item.patente_primaria = safe_text(payload.get("patente_primaria"), upper=True)
    item.patente_secundaria = safe_text(payload.get("patente_secundaria"), upper=True)
    item.bitren = safe_text(payload.get("bitren"), upper=True)
    item.tipo_unidad = safe_text(payload.get("tipo_unidad"))
    item.dni = safe_text(payload.get("dni"))
    item.nombre = safe_text(payload.get("nombre"))
    item.transporte = safe_text(payload.get("transporte"))
    item.base_descripcion = safe_text(payload.get("base_descripcion"))
    item.motivo = safe_text(payload.get("motivo"))
    item.evidence_original_name = evidence_original_name
    item.evidence_stored_name = evidence_stored_name
    item.ip_origen = ip_origen
    item.equipo = safe_text(payload.get("equipo"))
    item.usuario_sistema = safe_text(payload.get("usuario_sistema"))

    db.session.add(item)
    db.session.commit()

    mail_sent, mail_error = send_notification_email(item, requester_email=requester_email)
    if not mail_sent:
        add_audit(
            principal["username"],
            "MAIL_NOTIFICATION_ERROR",
            str(item.id),
            safe_text(mail_error)[:180] if mail_error else "Error de envio desconocido",
        )
        db.session.commit()

    return jsonify(
        {
            "message": "Solicitud registrada",
            "data": serialize_request(item),
            "mail_notification": {
                "sent": bool(mail_sent),
                "error": mail_error,
            },
        }
    ), 201


@app.route("/api/evidence/<int:request_id>", methods=["GET"])
def get_evidence_file(request_id: int):
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_view_records")
    if error_response:
        return error_response, status_code

    item = BlockRequest.query.get(request_id)
    if not item or not item.evidence_stored_name:
        return jsonify({"error": "Prueba no encontrada"}), 404

    file_path = os.path.join(EVIDENCE_DIR, item.evidence_stored_name)
    if not os.path.exists(file_path):
        return jsonify({"error": "Archivo de prueba no encontrado"}), 404

    return send_file(
        file_path,
        as_attachment=True,
        download_name=item.evidence_original_name or item.evidence_stored_name,
    )


@app.route("/api/requests", methods=["GET"])
def list_requests():
    principal, error_response, status_code = authenticate_operation_session()
    if error_response:
        return error_response, status_code

    limit = min(int(request.args.get("limit", 50)), 500)
    items = BlockRequest.query.order_by(desc(BlockRequest.created_at_system)).limit(limit).all()
    return jsonify([serialize_request(i) for i in items])


@app.route("/api/autocomplete", methods=["GET"])
def autocomplete():
    principal, error_response, status_code = authenticate_operation_session()
    if error_response:
        return error_response, status_code

    field = request.args.get("field", "")
    query_text = safe_text(request.args.get("q", ""))

    allowed = {
        "patente_primaria": BlockRequest.patente_primaria,
        "patente_secundaria": BlockRequest.patente_secundaria,
        "bitren": BlockRequest.bitren,
        "tipo_unidad": BlockRequest.tipo_unidad,
        "dni": BlockRequest.dni,
        "nombre": BlockRequest.nombre,
        "transporte": BlockRequest.transporte,
    }

    if field not in allowed:
        return jsonify({"error": "Campo no permitido"}), 400

    column = allowed[field]
    query = db.session.query(column).distinct().order_by(column)
    if query_text:
        query = query.filter(func.lower(column).like(f"%{query_text.lower()}%"))

    values = [row[0] for row in query.limit(12).all() if row[0]]
    return jsonify(values)


@app.route("/api/lookup", methods=["GET"])
def lookup():
    principal, error_response, status_code = authenticate_operation_session()
    if error_response:
        return error_response, status_code

    patente = safe_text(request.args.get("patente_primaria", ""), upper=True)
    patente_sec = safe_text(request.args.get("patente_secundaria", ""), upper=True)
    bitren = safe_text(request.args.get("bitren", ""), upper=True)
    dni = safe_text(request.args.get("dni", ""))
    nombre = safe_text(request.args.get("nombre", ""))

    if not patente and not patente_sec and not bitren and not dni and not nombre:
        return jsonify({})

    filters = []
    if patente:
        filters.append(BlockRequest.patente_primaria == patente)
        filters.append(BlockRequest.patente_secundaria == patente)
        filters.append(BlockRequest.bitren == patente)
    if patente_sec:
        filters.append(BlockRequest.patente_primaria == patente_sec)
        filters.append(BlockRequest.patente_secundaria == patente_sec)
        filters.append(BlockRequest.bitren == patente_sec)
    if bitren:
        filters.append(BlockRequest.patente_primaria == bitren)
        filters.append(BlockRequest.patente_secundaria == bitren)
        filters.append(BlockRequest.bitren == bitren)
    if dni:
        filters.append(BlockRequest.dni == dni)
    if nombre:
        filters.append(func.lower(BlockRequest.nombre) == nombre.lower())

    item = (
        BlockRequest.query.filter(or_(*filters))
        .order_by(desc(BlockRequest.created_at_system))
        .first()
    )

    excel_transport, matched_plate = lookup_transport_from_reference(patente, patente_sec, bitren)
    excel_unit_type, excel_unit_type_detail = lookup_unit_type_from_reference(patente, patente_sec, bitren)

    if not item:
        if excel_transport or excel_unit_type:
            return jsonify(
                {
                    "transporte": excel_transport,
                    "tipo_unidad": excel_unit_type,
                    "unit_type_lookup_detail": excel_unit_type_detail,
                    "transport_lookup_source": "excel",
                    "transport_lookup_plate": matched_plate,
                }
            )
        return jsonify({})

    return jsonify(
        {
            "id": item.id,
            "action_type": item.action_type,
            "tipo_unidad": excel_unit_type or item.tipo_unidad,
            "unit_type_lookup_detail": excel_unit_type_detail,
            "patente_secundaria": item.patente_secundaria,
            "bitren": item.bitren,
            "nombre": item.nombre,
            "transporte": excel_transport or item.transporte,
            "transport_lookup_source": "excel" if excel_transport else "registro",
            "transport_lookup_plate": matched_plate if excel_transport else None,
            "dni": item.dni,
            "patente_primaria": item.patente_primaria,
            "base_descripcion": item.base_descripcion,
        }
    )


@app.route("/api/reference/bases", methods=["GET"])
def reference_bases():
    principal, error_response, status_code = authenticate_operation_session()
    if error_response:
        return error_response, status_code

    bases, _, _ = load_reference_data()
    return jsonify(bases)


@app.route("/api/reference/transport-lookup", methods=["GET"])
def reference_transport_lookup():
    principal, error_response, status_code = authenticate_operation_session()
    if error_response:
        return error_response, status_code

    patente_primaria = safe_text(request.args.get("patente_primaria", ""), upper=True)
    patente_secundaria = safe_text(request.args.get("patente_secundaria", ""), upper=True)
    bitren = safe_text(request.args.get("bitren", ""), upper=True)

    transport, matched_plate = lookup_transport_from_reference(
        patente_primaria,
        patente_secundaria,
        bitren,
    )

    return jsonify(
        {
            "transporte": transport,
            "matched_plate": matched_plate,
            "has_match": bool(transport),
        }
    )


@app.route("/api/search/records", methods=["GET"])
def search_records():
    principal, error_response, status_code = authenticate_operation_session()
    if error_response:
        return error_response, status_code

    q = safe_text(request.args.get("q", ""))
    if not q:
        return jsonify([])

    limit = min(int(request.args.get("limit", 20)), 100)
    like = f"%{q.lower()}%"

    items = (
        BlockRequest.query.filter(
            or_(
                func.lower(BlockRequest.patente_primaria).like(like),
                func.lower(BlockRequest.patente_secundaria).like(like),
                func.lower(BlockRequest.dni).like(like),
                func.lower(BlockRequest.nombre).like(like),
                func.lower(BlockRequest.carga_nro).like(like),
                func.lower(BlockRequest.motivo).like(like),
            )
        )
        .order_by(desc(BlockRequest.created_at_system))
        .limit(limit)
        .all()
    )
    return jsonify([serialize_request(i) for i in items])


@app.route("/api/export/excel", methods=["GET"])
def export_excel():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_export_excel")
    if error_response:
        return error_response, status_code

    rows = BlockRequest.query.order_by(desc(BlockRequest.created_at_system)).all()

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Bloqueos"

    ws.append(
        [
            "ID",
            "FechaHoraSistemaGMT-3",
            "FechaHoraEvento",
            "Accion",
            "IDRelacionado",
            "SolicitadoPor",
            "AutorizadoPor",
            "NroCarga",
            "PatentePrimaria",
            "PatenteSecundaria",
            "Bitren",
            "TipoUnidad",
            "DNI",
            "Nombre",
            "Transporte",
            "Base",
            "Motivo",
            "PruebaArchivo",
            "IPOrigen",
            "Equipo",
            "UsuarioSistema",
        ]
    )

    for item in rows:
        ws.append(
            [
                item.id,
                item.created_at_system.isoformat(),
                item.event_datetime.isoformat(),
                item.action_type,
                item.related_request_id,
                item.usuario_sistema,
                item.authorized_by_user,
                item.carga_nro,
                item.patente_primaria,
                item.patente_secundaria,
                item.bitren,
                item.tipo_unidad,
                item.dni,
                item.nombre,
                item.transporte,
                item.base_descripcion,
                item.motivo,
                item.evidence_original_name,
                item.ip_origen,
                item.equipo,
                item.usuario_sistema,
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"bloqueos_{now_gmt_minus_3().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/meta", methods=["GET"])
def meta():
    return jsonify(
        {
            "server_datetime_gmt_minus_3": now_gmt_minus_3().isoformat(),
            "db_engine": app.config["SQLALCHEMY_DATABASE_URI"],
            "admin_enabled": is_admin_enabled(),
            "unlock_roles": get_unlock_allowed_roles(),
        }
    )


@app.route("/api/admin/status", methods=["GET"])
def admin_status():
    return jsonify({"admin_enabled": is_admin_enabled(), "users_enabled": has_user_auth()})


@app.route("/api/admin/verify", methods=["POST"])
def admin_verify():
    payload = request.get_json(silent=True) or {}
    admin_user = safe_text(payload.get("admin_user"))
    admin_key = safe_text(payload.get("admin_key"))

    if not LEGACY_ONLY_ADMIN:
        user = verify_user_credentials(admin_user, admin_key)
        if user:
            password_issue = get_user_password_issue(user)
            role_permissions = get_permissions_for_role(user.role)
            if not role_permissions.get("can_access_admin_panel", False):
                return jsonify({"error": "Sin permisos para acceder al panel admin"}), 403
            add_audit(user.username, "LOGIN_PANEL", user.username, f"role={user.role}")
            set_session_auth(user.username, user.role, "user")
            db.session.commit()
            return jsonify(
                {
                    "message": "Admin verificado",
                    "username": user.username,
                    "role": user.role,
                    "permissions": role_permissions,
                    "roles": sorted(ROLE_TYPES),
                    "can_unlock": bool(role_permissions.get("can_unlock_requests", False)),
                    "must_change_password": password_issue is not None,
                    "password_issue": password_issue,
                }
            )

    if can_use_legacy_admin(admin_user) and verify_admin_key(admin_key):
        role_permissions = get_permissions_for_role("ADMIN")
        set_session_auth("legacy_admin", "ADMIN", "legacy")
        return jsonify(
            {
                "message": "Admin verificado",
                "username": "legacy_admin",
                "role": "ADMIN",
                "permissions": role_permissions,
                "roles": sorted(ROLE_TYPES),
                "can_unlock": bool(role_permissions.get("can_unlock_requests", False)),
                "must_change_password": False,
                "password_issue": None,
            }
        )

    return jsonify({"error": "Credenciales admin invalidas"}), 403


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    clear_session_auth()
    return jsonify({"message": "Sesion cerrada"})


@app.route("/api/admin/requests", methods=["GET"])
def admin_list_requests():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_view_records")
    if error_response:
        return error_response, status_code

    limit = min(int(request.args.get("limit", 50)), 500)
    q = safe_text(request.args.get("q", ""))

    query = BlockRequest.query
    if q:
        like = f"%{q.lower()}%"
        query = query.filter(
            or_(
                func.lower(BlockRequest.patente_primaria).like(like),
                func.lower(BlockRequest.patente_secundaria).like(like),
                func.lower(BlockRequest.dni).like(like),
                func.lower(BlockRequest.nombre).like(like),
                func.lower(BlockRequest.carga_nro).like(like),
                func.lower(BlockRequest.motivo).like(like),
                func.lower(BlockRequest.action_type).like(like),
            )
        )

    items = query.order_by(desc(BlockRequest.created_at_system)).limit(limit).all()
    return jsonify([serialize_request(i) for i in items])


@app.route("/api/admin/requests/<int:request_id>", methods=["DELETE"])
def admin_delete_request(request_id: int):
    principal, error_response, status_code = authenticate_admin_request(required_roles={"ADMIN"})
    if error_response:
        return error_response, status_code
    assert principal is not None

    item = BlockRequest.query.get(request_id)
    if not item:
        return jsonify({"error": "Registro no encontrado"}), 404

    evidence_path = None
    if item.evidence_stored_name:
        evidence_path = os.path.join(EVIDENCE_DIR, item.evidence_stored_name)

    db.session.delete(item)
    add_audit(principal["username"], "DELETE_REQUEST", str(request_id), f"action={item.action_type}")
    db.session.commit()

    if evidence_path and os.path.exists(evidence_path):
        try:
            os.remove(evidence_path)
        except OSError:
            pass

    return jsonify({"message": "Registro eliminado", "id": request_id})


@app.route("/api/admin/users", methods=["GET"])
def admin_users_list():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_users")
    if error_response:
        return error_response, status_code

    users = AppUser.query.order_by(func.lower(AppUser.username)).all()
    return jsonify([serialize_user(user) for user in users])


@app.route("/api/admin/users", methods=["POST"])
def admin_users_create():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_users")
    if error_response:
        return error_response, status_code
    assert principal is not None

    payload = request.get_json(silent=True) or {}
    username = safe_text(payload.get("username"))
    role = safe_text(payload.get("role"), upper=True)
    key_plain = safe_text(payload.get("key_plain"))

    if not username or not key_plain:
        return jsonify({"error": "Usuario y clave son obligatorios"}), 400

    if role not in ROLE_TYPES:
        return jsonify({"error": "Rol invalido"}), 400

    policy = get_password_policy()
    policy_error = validate_password_strength(key_plain, policy)
    if policy_error:
        return jsonify({"error": policy_error}), 400

    if verify_admin_key(key_plain):
        return jsonify({"error": "La clave no puede coincidir con la admin key legacy"}), 400

    exists = AppUser.query.filter(func.lower(AppUser.username) == username.lower()).first()
    if exists:
        return jsonify({"error": "El usuario ya existe"}), 409

    now = now_gmt_minus_3()
    user = AppUser()
    user.username = username
    user.role = role
    user.key_hash = hash_key(key_plain)
    user.active = True
    user.must_change_password = True
    user.last_password_change_at = now
    user.created_at = now
    user.updated_at = now
    db.session.add(user)
    add_audit(principal["username"], "CREATE_USER", user.username, f"role={role}")
    db.session.commit()

    return jsonify({"message": "Usuario creado", "user": serialize_user(user)}), 201


@app.route("/api/admin/users/<int:user_id>", methods=["PATCH"])
def admin_users_update(user_id: int):
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_users")
    if error_response:
        return error_response, status_code
    assert principal is not None

    payload = request.get_json(silent=True) or {}
    user = AppUser.query.get(user_id)
    if not user:
        return jsonify({"error": "Usuario no encontrado"}), 404

    role = payload.get("role")
    if role is not None:
        role = safe_text(role, upper=True)
        if role not in ROLE_TYPES:
            return jsonify({"error": "Rol invalido"}), 400
        user.role = role
        add_audit(principal["username"], "UPDATE_ROLE", user.username, f"role={role}")

    active = payload.get("active")
    if active is not None:
        user.active = bool(active)
        add_audit(principal["username"], "UPDATE_ACTIVE", user.username, f"active={user.active}")

    user.updated_at = now_gmt_minus_3()
    db.session.commit()
    return jsonify({"message": "Usuario actualizado", "user": serialize_user(user)})


@app.route("/api/admin/users/<int:user_id>/reset-key", methods=["POST"])
def admin_users_reset_key(user_id: int):
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_users")
    if error_response:
        return error_response, status_code
    assert principal is not None

    payload = request.get_json(silent=True) or {}
    key_plain = safe_text(payload.get("key_plain"))
    if not key_plain:
        return jsonify({"error": "Nueva clave obligatoria"}), 400

    policy = get_password_policy()
    policy_error = validate_password_strength(key_plain, policy)
    if policy_error:
        return jsonify({"error": policy_error}), 400

    if verify_admin_key(key_plain):
        return jsonify({"error": "La clave no puede coincidir con la admin key legacy"}), 400

    user = AppUser.query.get(user_id)
    if not user:
        return jsonify({"error": "Usuario no encontrado"}), 404

    user.key_hash = hash_key(key_plain)
    user.must_change_password = True
    user.last_password_change_at = now_gmt_minus_3()
    user.updated_at = now_gmt_minus_3()
    add_audit(principal["username"], "RESET_HASH", user.username, "must_change_password=true")
    db.session.commit()

    return jsonify({"message": "Hash reseteado", "user": serialize_user(user)})


@app.route("/api/admin/users/force-password-rotation", methods=["POST"])
def admin_users_force_password_rotation():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_users")
    if error_response:
        return error_response, status_code
    assert principal is not None

    payload = request.get_json(silent=True) or {}
    legacy_only = bool(payload.get("legacy_only", True))
    include_inactive = bool(payload.get("include_inactive", False))

    query = AppUser.query
    if not include_inactive:
        query = query.filter(AppUser.active == True)

    users = query.all()
    updated = []

    for user in users:
        if user.username == principal["username"]:
            continue
        if legacy_only and get_hash_scheme(user.key_hash) != "legacy_sha256":
            continue

        user.must_change_password = True
        user.updated_at = now_gmt_minus_3()
        updated.append(user.username)

    if updated:
        add_audit(
            principal["username"],
            "FORCE_PASSWORD_ROTATION",
            "users",
            f"count={len(updated)},legacy_only={int(legacy_only)},include_inactive={int(include_inactive)}",
        )
        db.session.commit()

    return jsonify(
        {
            "message": "Recambio aplicado",
            "updated_count": len(updated),
            "updated_users": updated,
            "legacy_only": legacy_only,
            "include_inactive": include_inactive,
        }
    )


@app.route("/api/admin/change-own-key", methods=["POST"])
def admin_change_own_key():
    payload = request.get_json(silent=True) or {}
    username = safe_text(payload.get("username"))
    current_key = safe_text(payload.get("current_key"))
    new_key = safe_text(payload.get("new_key"))

    if not username or not current_key or not new_key:
        return jsonify({"error": "Usuario, clave actual y nueva clave son obligatorios"}), 400

    policy = get_password_policy()
    policy_error = validate_password_strength(new_key, policy)
    if policy_error:
        return jsonify({"error": policy_error}), 400

    if verify_admin_key(new_key):
        return jsonify({"error": "La nueva clave no puede coincidir con la admin key legacy"}), 400

    user = verify_user_credentials(username, current_key)
    if not user:
        return jsonify({"error": "Credenciales invalidas"}), 403

    user.key_hash = hash_key(new_key)
    user.must_change_password = False
    user.last_password_change_at = now_gmt_minus_3()
    user.updated_at = now_gmt_minus_3()
    add_audit(user.username, "CHANGE_OWN_KEY", user.username, "must_change_password=false")
    db.session.commit()

    return jsonify({"message": "Clave actualizada", "user": serialize_user(user)})


@app.route("/api/admin/audit", methods=["GET"])
def admin_audit_list():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_view_audit")
    if error_response:
        return error_response, status_code

    limit = min(int(request.args.get("limit", 100)), 500)
    logs = UserAuditLog.query.order_by(desc(UserAuditLog.created_at)).limit(limit).all()
    return jsonify([serialize_audit(item) for item in logs])


@app.route("/api/admin/users/password-health", methods=["GET"])
def admin_users_password_health():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_view_password_health")
    if error_response:
        return error_response, status_code

    threshold_days = max(0, min(int(request.args.get("days", 7)), 3650))
    policy = get_password_policy()

    users = AppUser.query.filter(AppUser.active == True).order_by(func.lower(AppUser.username)).all()

    expired = []
    expiring_soon = []
    ok = []

    for user in users:
        days_left = get_days_until_expiry(user)
        issue = get_user_password_issue(user)
        item = serialize_password_health_user(user)

        if issue in {"must_change_password", "password_expired"} or (days_left is not None and days_left < 0):
            expired.append(item)
        elif days_left is not None and days_left <= threshold_days:
            expiring_soon.append(item)
        else:
            ok.append(item)

    return jsonify(
        {
            "policy_expires_days": policy["expires_days"],
            "threshold_days": threshold_days,
            "expired_count": len(expired),
            "expiring_soon_count": len(expiring_soon),
            "ok_count": len(ok),
            "expired": expired,
            "expiring_soon": expiring_soon,
            "ok": ok,
        }
    )


@app.route("/api/admin/settings", methods=["GET"])
def admin_settings_get():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_policy")
    if error_response:
        return error_response, status_code
    return jsonify(get_password_policy())


@app.route("/api/admin/settings", methods=["PATCH"])
def admin_settings_update():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_policy")
    if error_response:
        return error_response, status_code
    assert principal is not None

    payload = request.get_json(silent=True) or {}

    try:
        min_length = max(6, int(payload.get("min_length", DEFAULT_PASSWORD_POLICY["min_length"])))
        expires_days = max(0, int(payload.get("expires_days", DEFAULT_PASSWORD_POLICY["expires_days"])))
    except (TypeError, ValueError):
        return jsonify({"error": "Valores numericos invalidos"}), 400

    require_uppercase = bool(payload.get("require_uppercase", DEFAULT_PASSWORD_POLICY["require_uppercase"]))
    require_lowercase = bool(payload.get("require_lowercase", DEFAULT_PASSWORD_POLICY["require_lowercase"]))
    require_digit = bool(payload.get("require_digit", DEFAULT_PASSWORD_POLICY["require_digit"]))
    require_symbol = bool(payload.get("require_symbol", DEFAULT_PASSWORD_POLICY["require_symbol"]))

    set_setting_value("password_min_length", str(min_length))
    set_setting_value("password_require_uppercase", "1" if require_uppercase else "0")
    set_setting_value("password_require_lowercase", "1" if require_lowercase else "0")
    set_setting_value("password_require_digit", "1" if require_digit else "0")
    set_setting_value("password_require_symbol", "1" if require_symbol else "0")
    set_setting_value("password_expires_days", str(expires_days))

    add_audit(
        principal["username"],
        "UPDATE_PASSWORD_POLICY",
        "password_policy",
        f"min={min_length},exp={expires_days},U={int(require_uppercase)},L={int(require_lowercase)},D={int(require_digit)},S={int(require_symbol)}",
    )
    db.session.commit()

    return jsonify({"message": "Politica actualizada", "policy": get_password_policy()})


@app.route("/api/admin/notifications", methods=["GET"])
def admin_notifications_get():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_policy")
    if error_response:
        return error_response, status_code
    return jsonify(get_notification_settings(include_secret=False))


@app.route("/api/admin/notifications", methods=["PATCH"])
def admin_notifications_update():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_policy")
    if error_response:
        return error_response, status_code
    assert principal is not None

    payload = request.get_json(silent=True) or {}
    ok, message = save_notification_settings(payload)
    if not ok:
        return jsonify({"error": message or "No se pudo guardar configuracion de notificaciones"}), 400

    add_audit(principal["username"], "UPDATE_NOTIFICATION_SETTINGS", "notifications", "Configuracion de mail actualizada")
    db.session.commit()
    return jsonify({"message": "Notificaciones actualizadas", "settings": get_notification_settings(include_secret=False)})


@app.route("/api/admin/notifications/test", methods=["POST"])
def admin_notifications_test_send():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_policy")
    if error_response:
        return error_response, status_code
    assert principal is not None

    payload = request.get_json(silent=True) or {}
    action_type = safe_text(payload.get("action_type"), upper=True) or "NOVEDAD"
    if action_type not in ACTION_TYPES:
        return jsonify({"error": "Tipo de accion invalido para prueba"}), 400

    item = BlockRequest()
    item.id = 0
    item.created_at_system = now_gmt_minus_3()
    item.event_datetime = now_gmt_minus_3()
    item.action_type = action_type
    item.related_request_id = None
    item.authorized_by_user = None
    item.carga_nro = "PRUEBA"
    item.patente_primaria = "TEST123"
    item.patente_secundaria = ""
    item.bitren = ""
    item.tipo_unidad = "Unidad de prueba"
    item.dni = "00000000"
    item.nombre = "Usuario de prueba"
    item.transporte = "Transporte de prueba"
    item.base_descripcion = "Base de prueba"
    item.motivo = "Correo de prueba desde configuracion de notificaciones"
    item.evidence_original_name = None
    item.evidence_stored_name = None
    item.ip_origen = "127.0.0.1"
    item.equipo = "panel-admin"
    item.usuario_sistema = principal["username"]

    sent, error_message = send_notification_email(item)
    if not sent:
        add_audit(principal["username"], "TEST_NOTIFICATION_EMAIL_ERROR", "notifications", safe_text(error_message)[:180] if error_message else "Error de envio")
        db.session.commit()
        return jsonify({"error": error_message or "No se pudo enviar mail de prueba"}), 400

    add_audit(principal["username"], "TEST_NOTIFICATION_EMAIL_OK", "notifications", f"action_type={action_type}")
    db.session.commit()
    return jsonify({"message": "Mail de prueba enviado"})


@app.route("/api/admin/permissions", methods=["GET"])
def admin_permissions_get():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_permissions")
    if error_response:
        return error_response, status_code

    return jsonify(
        {
            "roles": sorted(ROLE_TYPES),
            "permission_keys": sorted(PERMISSION_KEYS),
            "matrix": get_role_permissions_matrix(),
        }
    )


@app.route("/api/admin/permissions", methods=["PATCH"])
def admin_permissions_update():
    principal, error_response, status_code = authenticate_admin_request(required_roles=ROLE_TYPES, required_permission="can_manage_permissions")
    if error_response:
        return error_response, status_code
    assert principal is not None

    payload = request.get_json(silent=True) or {}
    matrix = payload.get("matrix")
    if not isinstance(matrix, dict):
        return jsonify({"error": "Payload invalido para permisos"}), 400

    save_role_permissions_matrix(matrix)
    add_audit(principal["username"], "UPDATE_ROLE_PERMISSIONS", "role_permissions", "Matriz de permisos actualizada")
    db.session.commit()

    return jsonify({"message": "Permisos actualizados", "matrix": get_role_permissions_matrix()})


with app.app_context():
    db.create_all()
    ensure_block_request_schema()
    ensure_user_schema()
    ensure_settings_defaults()


if __name__ == "__main__":
    ssl_cert = os.getenv("SSL_CERT_FILE", "").strip()
    ssl_key = os.getenv("SSL_KEY_FILE", "").strip()
    debug_mode = os.getenv("FLASK_DEBUG", "0").strip().lower() in {"1", "true", "yes", "on"}
    ssl_context = None
    if ssl_cert and ssl_key and os.path.exists(ssl_cert) and os.path.exists(ssl_key):
        ssl_context = (ssl_cert, ssl_key)

    app.run(host="0.0.0.0", port=5000, debug=debug_mode, ssl_context=ssl_context)

import os
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
TARGET_FILE = os.path.join(BASE_DIR, "data.xlsx")
BACKUP_DIR = os.path.join(BASE_DIR, "instance", "reference_backups")

REQUIRED_SHEETS = {
    "Plantas": ["Descripcion de Base"],
    "Transportes": ["Inicial de equipo", "Transportista", "ID de tipo de equipo"],
}


def normalize_header(value: str) -> str:
    text = (value or "").strip().lower()
    replacements = {
        "a": ["a", "á", "à", "ä", "â"],
        "e": ["e", "é", "è", "ë", "ê"],
        "i": ["i", "í", "ì", "ï", "î"],
        "o": ["o", "ó", "ò", "ö", "ô"],
        "u": ["u", "ú", "ù", "ü", "û"],
        "n": ["n", "ñ"],
    }
    for plain, variants in replacements.items():
        for variant in variants:
            text = text.replace(variant, plain)
    return " ".join(text.split())


def validate_excel(path: str) -> tuple[bool, str]:
    if not os.path.exists(path):
        return False, f"No existe el archivo: {path}"

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return False, f"No se pudo abrir el Excel: {exc}"

    try:
        sheet_names = set(wb.sheetnames)
        for sheet_name, required_headers in REQUIRED_SHEETS.items():
            if sheet_name not in sheet_names:
                return False, f"Falta la hoja requerida: {sheet_name}"

            ws = wb[sheet_name]
            headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
            normalized_headers = {normalize_header(str(item or "")) for item in headers}

            for required in required_headers:
                if normalize_header(required) not in normalized_headers:
                    return False, f"Falta columna '{required}' en hoja {sheet_name}"
    finally:
        wb.close()

    return True, "OK"


def main() -> int:
    if len(sys.argv) < 2:
        print("Uso: python update_reference_data.py <ruta_excel_nuevo>")
        return 1

    source_path = os.path.abspath(sys.argv[1])
    ok, message = validate_excel(source_path)
    if not ok:
        print(f"ERROR: {message}")
        return 1

    os.makedirs(BACKUP_DIR, exist_ok=True)

    if os.path.exists(TARGET_FILE):
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"data_{stamp}.xlsx"
        backup_path = os.path.join(BACKUP_DIR, backup_name)
        shutil.copy2(TARGET_FILE, backup_path)
        print(f"Backup creado: {backup_path}")

    shutil.copy2(source_path, TARGET_FILE)
    print(f"OK: Referencia actualizada en {TARGET_FILE}")
    print("Reinicia la app para tomar los cambios inmediatamente.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
